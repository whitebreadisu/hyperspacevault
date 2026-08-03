"""BL-186: app/services/swunlimiteddb_import.py -- sw-unlimited-db.com
collection-export (XLSX) import adapter. Pure unit tests only (set-code
table, normalization, column-family resolution, melt/positive-only
semantics, format detection) -- no LIVE Postgres connection is ever used,
mirroring test_swudb_import.py's posture (DATABASE_URL/APP_DATABASE_URL
still need to be SET as env vars -- app/tests/conftest.py's import chain
reads them at collection time -- but nothing here actually queries the
DB). The melt tests build real in-memory XLSX bytes via openpyxl and
monkeypatch the one DB touchpoint (repositories/inventory_import.
get_base_card_variants_by_set_and_number), so they exercise the actual
parse_swunlimiteddb_xlsx code path without a live database. DB-backed
resolution/API-level coverage (parse_swunlimiteddb_xlsx against real
card_variants rows, POST /api/inventory/import end to end) lives in
test_inventory_import_api.py alongside the rest of the import API tests,
reusing its existing tenant/catalog fixture plumbing.

Run:
    docker compose exec backend pytest app/tests/test_swunlimiteddb_import.py -v
"""

import io

import openpyxl
import pytest

from app.models.card_variant import CardVariant
from app.repositories import inventory_import as inventory_import_repo
from app.services import swunlimiteddb_import as swi


def _variant(variant_type: str, swuapi_id: str, source_set_code="SOR", card_number="1"):
    """A bare (unpersisted) CardVariant plus the (name, subtitle, type)
    triple _resolve_column expects -- VariantMatch's shape, built without
    touching the DB (matches test_swudb_import.py's own `_variant` helper)."""
    cv = CardVariant(
        variant_type=variant_type,
        swuapi_id=swuapi_id,
        source_set_code=source_set_code,
        card_number=card_number,
    )
    return (cv, "Test Card", None, "Unit")


def _xlsx_bytes(header: list[str], rows: list[list]) -> bytes:
    """Builds a real single-sheet ("Data") XLSX workbook in memory --
    exercises openpyxl's actual read path (not a fixture stub), matching
    the definition doc's own "read via openpyxl" method."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = swi.DATA_SHEET_NAME
    ws.append(header)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestSetCodeMap:
    def test_direct_base_set_codes(self):
        for code in (
            "SOR",
            "SHD",
            "TWI",
            "JTL",
            "LAW",
            "LOF",
            "SEC",
            "ASH",
            "IBH",
            "TS26",
        ):
            assert swi.SET_CODE_MAP[code] == code

    def test_hmw_absent(self):
        """HMW (BL-186 definition doc §3.3/§10): a genuinely unreleased/
        preview set per the owner's read -- deliberately NOT in the table,
        the owner's standing rule being this catalog only maps
        swuapi-sourced content. Never guessed."""
        assert "HMW" not in swi.SET_CODE_MAP


class TestNormalizeNumber:
    def test_strips_three_digit_padding(self):
        assert swi._normalize_number("001") == "1"

    def test_strips_two_digit_padding(self):
        assert swi._normalize_number("08") == "8"

    def test_no_padding_is_a_no_op(self):
        assert swi._normalize_number("1127") == "1127"

    def test_mixed_padding_within_one_set_still_normalizes(self):
        """§2.2: unlike SWUDB, padding is inconsistent WITHIN a single set
        (e.g. sec) -- the recipe is per-cell, not per-set, so it handles
        both shapes identically."""
        assert swi._normalize_number("002") == "2"
        assert swi._normalize_number("101") == "101"

    def test_token_t_prefix_stripped_then_normalized(self):
        assert swi._normalize_number("T02") == "2"
        assert swi._normalize_number("T2") == "2"

    def test_non_numeric_falls_back_to_stripped(self):
        assert swi._normalize_number(" abc ") == "abc"


class TestCellQuantity:
    def test_blank_is_no_claim(self):
        assert swi._cell_quantity(None) is None

    def test_literal_zero_is_ignored_as_no_claim(self):
        """§2.3/§10 (owner-locked): positive-only semantics -- a literal 0
        cell, should one ever appear, is treated identically to blank."""
        assert swi._cell_quantity(0) is None

    def test_positive_int_is_the_claim(self):
        assert swi._cell_quantity(3) == 3

    def test_negative_is_ignored(self):
        assert swi._cell_quantity(-1) is None

    def test_non_numeric_is_ignored(self):
        assert swi._cell_quantity("not-a-number") is None


class TestResolveColumnCleanFamilies:
    def test_normal_resolves_standard(self):
        variants = [_variant("Standard", "uuid-std")]
        outcome = swi._resolve_column(variants, "Normal", "SOR")
        assert outcome.swuapi_uuid == "uuid-std"
        assert outcome.reason is None

    def test_foil_resolves_standard_foil(self):
        variants = [
            _variant("Standard", "uuid-std"),
            _variant("Standard Foil", "uuid-foil"),
        ]
        outcome = swi._resolve_column(variants, "Foil", "SOR")
        assert outcome.swuapi_uuid == "uuid-foil"

    def test_foil_and_hyperspace_resolves_hyperspace_foil(self):
        """§4 row 32: sw-unlimited-db's "Foil & Hyperspace" column name
        order is reversed vs. our "Hyperspace Foil" variant_type -- same
        semantics."""
        variants = [_variant("Hyperspace Foil", "uuid-hf")]
        outcome = swi._resolve_column(variants, "Foil & Hyperspace", "SOR")
        assert outcome.swuapi_uuid == "uuid-hf"

    def test_showcase_resolves(self):
        variants = [_variant("Showcase", "uuid-sc")]
        outcome = swi._resolve_column(variants, "Showcase", "SOR")
        assert outcome.swuapi_uuid == "uuid-sc"

    def test_zero_candidates_is_unknown_variant_for_column(self):
        outcome = swi._resolve_column([], "Normal", "SOR")
        assert outcome.swuapi_uuid is None
        assert outcome.reason == "unknown_variant_for_column"
        assert outcome.candidates == []


class TestResolveColumnWeeklyPlay:
    def test_organized_play_resolves_via_weekly_play_type(self):
        variants = [
            _variant("Standard", "uuid-std"),
            _variant("Weekly Play", "uuid-wp", source_set_code="JTLP"),
        ]
        outcome = swi._resolve_column(variants, "Organized Play", "JTL")
        assert outcome.swuapi_uuid == "uuid-wp"

    def test_organized_play_resolves_via_early_era_container_set(self):
        """§10 owner finding (BL-183): an early-era root-coded WP variant
        can carry a DIFFERENT variant_type (here "Hyperspace") while living
        in the `<set>P` container -- still resolves via the source_set_code
        leg of the family predicate, not the variant_type leg."""
        variants = [
            _variant("Standard", "uuid-std", source_set_code="SOR"),
            _variant("Hyperspace", "uuid-wp-early", source_set_code="SORP"),
        ]
        outcome = swi._resolve_column(variants, "Organized Play", "SOR")
        assert outcome.swuapi_uuid == "uuid-wp-early"

    def test_no_weekly_play_family_member_is_unresolved(self):
        variants = [_variant("Standard", "uuid-std")]
        outcome = swi._resolve_column(variants, "Organized Play", "SOR")
        assert outcome.swuapi_uuid is None
        assert outcome.reason == "unknown_variant_for_column"


class TestResolveColumnPrerelease:
    def test_prefers_exact_promo_over_judge(self):
        """§5.4/§10: when both Prerelease Promo and Prerelease Judge exist
        on the same base card, Prerelease Promo resolves to the exact
        match rather than going ambiguous."""
        variants = [
            _variant("Prerelease Promo", "uuid-promo"),
            _variant("Prerelease Judge", "uuid-judge"),
        ]
        outcome = swi._resolve_column(variants, "Prerelease Promo", "SOR")
        assert outcome.swuapi_uuid == "uuid-promo"

    def test_sole_prerelease_judge_still_resolves_when_no_promo_exists(self):
        """No exact "Prerelease Promo" match -- the "Prerelease" family
        still narrows to whatever IS there when it's a single candidate."""
        variants = [_variant("Prerelease Judge", "uuid-judge")]
        outcome = swi._resolve_column(variants, "Prerelease Promo", "SOR")
        assert outcome.swuapi_uuid == "uuid-judge"


class TestResolveColumnHomeSetPreference:
    def test_prefers_the_home_set_printing_among_a_tied_family(self):
        """§10 step 2: two Weekly Play candidates on one base card, only
        one matching the row's own mapped set -- disambiguates by home-set
        preference rather than going ambiguous."""
        variants = [
            _variant("Weekly Play", "uuid-home", source_set_code="LAW"),
            _variant("Weekly Play", "uuid-other", source_set_code="LAWP"),
        ]
        outcome = swi._resolve_column(variants, "Organized Play", "LAW")
        assert outcome.swuapi_uuid == "uuid-home"

    def test_still_ambiguous_when_no_home_set_member_disambiguates(self):
        """The Serialized Prestige color-trio shape (BL-186 doc §0.6): all
        three candidates share the same source_set_code as the mapped set,
        so the home-set filter can't narrow -- irreducibly ambiguous,
        routed to the EXISTING ambiguous_triple reason (same locked
        decision BL-185 made: no new UX)."""
        variants = [
            _variant("Serialized Prestige", "uuid-c1", source_set_code="SEC"),
            _variant("Serialized Prestige", "uuid-c2", source_set_code="SEC"),
            _variant("Serialized Prestige", "uuid-c3", source_set_code="SEC"),
        ]
        outcome = swi._resolve_column(variants, "Serialized Prestige", "SEC")
        assert outcome.swuapi_uuid is None
        assert outcome.reason == "ambiguous_triple"
        assert outcome.candidates == ["uuid-c1", "uuid-c2", "uuid-c3"]


class TestResolveColumnBaseCardFirst:
    def test_resolves_despite_different_own_number_and_source_set(self):
        """§10 (load-bearing owner finding): base_card_variants is already
        scoped to ONE base card (the caller's job, via base_card_id) -- a
        family member with a totally different own card_number/
        source_set_code from the row's own (mapped_set, number) still
        resolves, because selection never re-derives identity from those
        fields at all."""
        variants = [
            _variant("Standard", "uuid-std", source_set_code="SOR", card_number="9836"),
            _variant(
                "Prerelease Promo",
                "uuid-prerelease",
                source_set_code="P25",
                card_number="9736",
            ),
        ]
        outcome = swi._resolve_column(variants, "Prerelease Promo", "SOR")
        assert outcome.swuapi_uuid == "uuid-prerelease"


class TestResolveColumnUnmapped:
    def test_event_exclusive_is_unmapped_column_even_with_a_plausible_candidate(self):
        """§6/§10 (owner-locked): Event Exclusive's single observed
        real-world use is a card-level anomaly -- unmapped_column in v1,
        even when a variant that LOOKS right (e.g. an Event Exclusive-
        named variant_type) exists on the base card."""
        variants = [_variant("Standard", "uuid-std")]
        outcome = swi._resolve_column(variants, "Event Exclusive", "SOR")
        assert outcome.swuapi_uuid is None
        assert outcome.reason == "unmapped_column"

    @pytest.mark.parametrize(
        "column",
        [
            "Organized Play Foil",
            "Top 8",
            "Top 4",
            "Finalist",
            "Champion",
            "Participation",
            "Judge",
            "Price wall",
            "Event pack",
            "Day two reward",
        ],
    )
    def test_dead_columns_are_unmapped_column(self, column):
        outcome = swi._resolve_column([], column, "SOR")
        assert outcome.reason == "unmapped_column"


class TestFormatDetectionXlsx:
    """The router's XLSX sniff (app/routers/inventory.py's _looks_like_xlsx)
    -- pure function, no DB."""

    def test_xlsx_extension_detected_even_with_non_xlsx_bytes(self):
        from app.routers.inventory import _looks_like_xlsx

        assert _looks_like_xlsx("collection.xlsx", b"not really xlsx bytes") is True

    def test_magic_prefix_detected_without_an_xlsx_extension(self):
        from app.routers.inventory import _looks_like_xlsx

        assert _looks_like_xlsx("upload", b"PK\x03\x04rest-of-the-zip") is True

    def test_case_insensitive_extension(self):
        from app.routers.inventory import _looks_like_xlsx

        assert _looks_like_xlsx("Collection.XLSX", b"") is True

    def test_plain_csv_bytes_not_detected(self):
        from app.routers.inventory import _looks_like_xlsx

        assert _looks_like_xlsx("upload.csv", b"Set,CardNumber,Count,IsFoil\n") is False

    def test_pk_looking_text_without_the_full_magic_is_not_detected(self):
        """Cheap edge case: text that merely starts with the letters "PK"
        (not the real 4-byte ZIP local-file-header signature) must not be
        misdetected as XLSX."""
        from app.routers.inventory import _looks_like_xlsx

        assert _looks_like_xlsx("upload.csv", b"PK,CardNumber,Count,IsFoil\n") is False


class TestParseMeltAndPositiveOnly:
    """End-to-end through parse_swunlimiteddb_xlsx with real XLSX bytes,
    monkeypatching the one DB call so these stay DATABASE_URL-free -- the
    resolution outcome is irrelevant here (every pair maps to a fixture
    with an empty variant list), only the MELT shape (how many candidate
    rows a wide row produces, in what order, with what quantities) and the
    positive-only semantics matter."""

    def _parse_with_empty_catalog(self, monkeypatch, header, rows):
        monkeypatch.setattr(
            inventory_import_repo,
            "get_base_card_variants_by_set_and_number",
            lambda db, pairs: {},
        )
        return swi.parse_swunlimiteddb_xlsx(_xlsx_bytes(header, rows), db=None)

    def test_vader_style_row_melts_into_multiple_candidates(self, monkeypatch):
        """§0.1/§4 row 1: one row with Normal/Showcase/Prerelease Promo all
        populated melts into 3 separate report rows, not 1."""
        header = [
            "Set",
            "Base card id",
            "Name",
            "Normal",
            "Showcase",
            "Prerelease Promo",
        ]
        rows = [["sor", "010", "Darth Vader", 1, 1, 1]]
        result = self._parse_with_empty_catalog(monkeypatch, header, rows)
        assert len(result.rows) == 3
        assert {r.quantity for r in result.rows} == {1}

    def test_blank_cells_produce_no_candidate(self, monkeypatch):
        header = ["Set", "Base card id", "Name", "Normal", "Foil", "Showcase"]
        rows = [["sor", "9841", "BL186 Test", 2, None, None]]
        result = self._parse_with_empty_catalog(monkeypatch, header, rows)
        assert len(result.rows) == 1
        assert result.rows[0].quantity == 2

    def test_literal_zero_cell_produces_no_candidate(self, monkeypatch):
        """§2.3/§10 (owner-locked): a literal 0 -- distinct from blank --
        is still ignored as no-claim, never treated as a removal/zero
        write."""
        header = ["Set", "Base card id", "Name", "Normal", "Showcase"]
        rows = [["sor", "9841", "BL186 Test", 2, 0]]
        result = self._parse_with_empty_catalog(monkeypatch, header, rows)
        assert len(result.rows) == 1
        assert result.rows[0].quantity == 2

    def test_row_with_no_positive_cells_melts_to_nothing(self, monkeypatch):
        header = ["Set", "Base card id", "Name", "Normal", "Foil"]
        rows = [["sor", "9841", "BL186 Test", None, 0]]
        result = self._parse_with_empty_catalog(monkeypatch, header, rows)
        assert result.rows == []

    def test_multiple_rows_each_melt_independently(self, monkeypatch):
        header = ["Set", "Base card id", "Name", "Normal", "Foil"]
        rows = [
            ["sor", "9841", "Row A", 3, None],
            ["shd", "9842", "Row B", None, 1],
        ]
        result = self._parse_with_empty_catalog(monkeypatch, header, rows)
        assert len(result.rows) == 2
        assert {r.quantity for r in result.rows} == {1, 3}


class TestParseDirectRefusals:
    """parse_swunlimiteddb_xlsx's own header/sheet checks -- belt-and-
    suspenders against the router's format-detection gate (which already
    guarantees XLSX-shaped bytes before dispatching here)."""

    def test_not_a_workbook_at_all_is_unparseable(self):
        with pytest.raises(swi.UnparseableFileError):
            swi.parse_swunlimiteddb_xlsx(b"not an xlsx file", db=None)

    def test_missing_data_sheet_is_unparseable(self):
        wb = openpyxl.Workbook()
        wb.active.title = "NotData"
        buf = io.BytesIO()
        wb.save(buf)
        with pytest.raises(swi.UnparseableFileError):
            swi.parse_swunlimiteddb_xlsx(buf.getvalue(), db=None)

    def test_header_missing_required_columns_is_unparseable(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = swi.DATA_SHEET_NAME
        ws.append(["Name", "Normal"])
        buf = io.BytesIO()
        wb.save(buf)
        with pytest.raises(swi.UnparseableFileError):
            swi.parse_swunlimiteddb_xlsx(buf.getvalue(), db=None)
