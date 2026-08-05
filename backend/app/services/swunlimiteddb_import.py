"""BL-186: sw-unlimited-db.com collection-export import adapter (XLSX).

sw-unlimited-db's "Grouped" export ("Detailed" 500s consistently, per the
BL-186 definition doc's owner session -- Grouped is the real, permanent
ceiling for this source) is a WIDE spreadsheet: one row per BASE card
(`Set`, `Base card id`, `Name`), followed by 21 fixed quantity columns, one
per named variant/finish/promo-tier bucket. A row's identity is the base
card; a printing's identity is `(Set, Base card id, <column name>)`, not a
per-row field the way SWUDB's `IsFoil`/`Stamp` worked (see swudb_import.py).
Parsing this format is therefore a MELT first (one wide row -> up to 21
candidate rows), not a straight row-by-row read.

Like swudb_import.parse_swudb_csv, this module resolves each candidate
against `card_variants` at PARSE time (there's no canonical (set_code,
card_number, variant_type) triple to hand off to inventory_import.py's own
resolution engine -- a column names a FAMILY of variant_types, not a single
one, and the row's own `Base card id` is the base card's number, not
necessarily any resolved variant's own card_number). A resolved candidate
carries only `swuapi_uuid` on its returned ParsedRow (set_code/card_number/
variant_type stay None) and flows through
`app.services.inventory_import.compute_import` completely UNCHANGED, via
that module's existing "uuid known -> resolved" step. An unresolved/
ambiguous candidate carries its verdict on ParsedRow.preresolved_reason/
preresolved_candidates instead, exactly like a SWUDB row -- see swudb_
import.py's module docstring and inventory_import.py's `_resolve_row` for
the shared contract both adapters stash into.

Resolution is BASE-CARD-FIRST (the BL-186 definition doc's §10 owner
finding, load-bearing): find the base card by (mapped_set,
normalized_number) -- base_cards joined via its set, NOT a variant's own
(source_set_code, card_number) -- then select among THAT CARD's variants by
the column's family. This is deliberately different from SWUDB's own
(set, card_number) variant lookup (swudb_import._resolve_candidates): a
promo/prerelease/Weekly-Play variant can carry a totally different own
number in a different source_set_code than its base card (the doc's worked
example: Vader's Prerelease Promo printing lands on the base sor/010 row,
not on any row keyed by its own number). See app/repositories/
inventory_import.py's get_base_card_variants_by_set_and_number for the bulk
fetch this resolves against.

The BL-186 definition doc (backlog-tracked analysis, not duplicated here)
is authoritative for every judgment call below; each is called out inline
as "§n" against that doc's section numbering.
"""

from __future__ import annotations

import io
from dataclasses import dataclass

import openpyxl
from sqlalchemy.orm import Session

from app.repositories import inventory_import as inventory_import_repo
from app.repositories.inventory_import import VariantMatch
from app.services.inventory_io import (
    ParsedRow,
    ParseNotes,
    ParseResult,
    UnparseableFileError,
    _merge_duplicates,
)

DATA_SHEET_NAME = "Data"

# §2.1: the row-identity columns, plus the 21 fixed quantity columns in
# their exported order. `Name` is descriptive only (never consulted for
# resolution, same posture as inventory_io's own name/subtitle fields).
IDENTITY_COLUMNS = ("Set", "Base card id", "Name")
QUANTITY_COLUMNS = (
    "Normal",
    "Foil",
    "Hyperspace",
    "Foil & Hyperspace",
    "Showcase",
    "Organized Play",
    "Event Exclusive",
    "Prerelease Promo",
    "Organized Play Foil",
    "Standard Prestige",
    "Foil Prestige",
    "Serialized Prestige",
    "Top 8",
    "Top 4",
    "Finalist",
    "Champion",
    "Participation",
    "Judge",
    "Price wall",
    "Event pack",
    "Day two reward",
)
KNOWN_COLUMNS = set(IDENTITY_COLUMNS) | set(QUANTITY_COLUMNS)
# §5.2/§10: header recognizability check -- Set and Base card id are the
# row's whole identity; Name/quantity columns are optional (a header with
# zero populated quantity columns is still a recognizable, just empty,
# export).
REQUIRED_COLUMNS = {"Set", "Base card id"}

# §3.1/§3.2/§10: the 10 base sets, case-fold (.upper()) before lookup --
# owner's standing rule (2026-08-03): this catalog maps only swuapi-sourced
# content, so ANYTHING else (incl. HMW, a genuinely unreleased/preview set
# per the owner's read of its card names) stays unmapped by design, never
# guessed. Extension policy mirrors swudb_import.SET_CODE_MAP: add a row
# only once a set is actually ingested from swuapi.
SET_CODE_MAP: dict[str, str] = {
    "SOR": "SOR",
    "SHD": "SHD",
    "TWI": "TWI",
    "JTL": "JTL",
    "LAW": "LAW",
    "LOF": "LOF",
    "SEC": "SEC",
    "ASH": "ASH",
    "IBH": "IBH",
    "TS26": "TS26",
}

# §5.4/§10: columns that resolve to a single, direct variant_type on the
# base card -- clean 1:1, no family ambiguity beyond the standard home-set/
# ambiguous disambiguation every column goes through.
_CLEAN_COLUMN_TO_TYPE: dict[str, str] = {
    "Normal": "Standard",
    "Foil": "Standard Foil",
    "Hyperspace": "Hyperspace",
    "Foil & Hyperspace": "Hyperspace Foil",
    "Showcase": "Showcase",
    "Standard Prestige": "Standard Prestige",
    "Foil Prestige": "Foil Prestige",
    "Serialized Prestige": "Serialized Prestige",
}

# §10 owner finding: Organized Play is the Weekly Play family -- either the
# variant's own type is literally "Weekly Play", OR it lives in the early-
# era root-coded WP container set (BL-183: `<mapped_set>P`, e.g. SORP/JTLP).
_WEEKLY_PLAY_COLUMN = "Organized Play"

# §5.4/§10: Prerelease Promo resolves within the base card's "Prerelease"-
# family variants, preferring the exact "Prerelease Promo" type when both
# Promo and Judge printings exist (a remaining multi-match is ambiguous).
_PRERELEASE_COLUMN = "Prerelease Promo"

# §6/§10 (owner-locked): columns that NEVER resolve in v1, surfaced as
# unmapped_column if ever positive rather than guessed. Event Exclusive's
# single observed real-world use (Vader's SORP_2) is a card-level anomaly,
# not a general rule -- no confident family mapping exists for it. The
# other 9 are the empirically-dead tournament/tier columns (0/2,255 rows in
# the owner's real export) plus Organized Play Foil and the three reward
# columns, none of which have ever been observed populated.
UNMAPPED_COLUMNS = frozenset(
    {
        "Event Exclusive",
        "Organized Play Foil",
        "Top 8",
        "Top 4",
        "Finalist",
        "Champion",
        "Participation",
        "Judge",
        "Price wall",
        "Event pack",
        "Day two reward",
    }
)


def _is_token_id(raw: str) -> bool:
    """BL-199: whether a raw `Base card id` names a TOKEN (T-prefixed,
    e.g. T1). The prefix is the row's only token signal, and it's decisive:
    because _normalize_number strips it before lookup, a token row and a
    real-card row would otherwise land on the SAME (set, number) pair --
    token base cards are numbered run-locally inside base sets with bare
    numbers (ASH 1 is both The Armorer and the Mandalorian token). The
    caller routes T-rows to the token side of the base-card lookup and
    bare rows to the non-token side (see get_base_card_variants_by_set_
    and_number's `tokens` param), so neither ever sees the other's
    candidates."""
    return raw.strip().startswith("T")


def _normalize_number(raw: str) -> str:
    """§2.2: strips a leading `T` (token ids) before the same int-cast/
    re-stringify recipe swudb_import._normalize_number uses -- resolves
    inconsistent padding (within AND across sets in this format) uniformly.
    A token's `T`-prefix is stripped entirely, matching how our own
    card_number field already stores tokens (BL-185 §3.4's P25_T002
    finding: the T-prefixed collector number and the plain numeric
    card_number are the same bare digits) -- tokenness itself is preserved
    by the caller via _is_token_id above (BL-199), not by this string.
    Falls back to the T-stripped, stripped string when it isn't purely
    numeric -- never raises; an unparseable number just fails to find any
    base card downstream (unknown_set_and_number)."""
    stripped = raw.strip()
    if stripped.startswith("T"):
        stripped = stripped[1:]
    try:
        return str(int(stripped))
    except ValueError:
        return stripped


@dataclass(frozen=True)
class _Outcome:
    swuapi_uuid: str | None
    reason: str | None
    candidates: list[str]


def _finish_family_resolution(family: list[VariantMatch], mapped_set: str) -> _Outcome:
    """§10 step 2: disambiguate a column's family of candidates (already
    narrowed to variants belonging to ONE base card) -- prefer the variant
    whose source_set_code equals the row's mapped set (the home-set
    printing); a remaining multi-match is the existing ambiguous_triple
    bucket (the Serialized Prestige family is the expected case, per the
    doc's §0.6/§9 Q3 census)."""
    if not family:
        return _Outcome(None, "unknown_variant_for_column", [])
    if len(family) == 1:
        return _Outcome(family[0][0].swuapi_id, None, [])
    home = [v for v in family if v[0].source_set_code == mapped_set]
    if len(home) == 1:
        return _Outcome(home[0][0].swuapi_id, None, [])
    candidates = sorted({v[0].swuapi_id for v in family})
    return _Outcome(None, "ambiguous_triple", candidates)


def _resolve_column(
    base_card_variants: list[VariantMatch], column: str, mapped_set: str
) -> _Outcome:
    """§5.4/§10: column -> family -> resolved/unresolved/ambiguous, applied
    to one base card's full variant list (already fetched base-card-first
    by the caller). Never touches the DB itself -- pure selection over the
    list the caller passed in."""
    if column in UNMAPPED_COLUMNS:
        # §6/§10 (owner-locked): no confident family mapping exists for
        # these columns -- surfaced as a problem row, never guessed, even
        # when the base card and a plausible-looking candidate both exist.
        return _Outcome(None, "unmapped_column", [])

    if column in _CLEAN_COLUMN_TO_TYPE:
        target_type = _CLEAN_COLUMN_TO_TYPE[column]
        family = [v for v in base_card_variants if v[0].variant_type == target_type]
        return _finish_family_resolution(family, mapped_set)

    if column == _WEEKLY_PLAY_COLUMN:
        wp_container = mapped_set + "P"
        family = [
            v
            for v in base_card_variants
            if v[0].variant_type == "Weekly Play"
            or v[0].source_set_code == wp_container
        ]
        return _finish_family_resolution(family, mapped_set)

    if column == _PRERELEASE_COLUMN:
        family = [v for v in base_card_variants if "Prerelease" in v[0].variant_type]
        exact = [v for v in family if v[0].variant_type == "Prerelease Promo"]
        if exact:
            family = exact
        return _finish_family_resolution(family, mapped_set)

    # Defensive only: every QUANTITY_COLUMNS entry is handled by one of the
    # branches above -- reachable only if this function is called with a
    # column outside that vocabulary, which the parser below never does.
    return _Outcome(None, "unmapped_column", [])


@dataclass(frozen=True)
class _Candidate:
    """One melted (row, column) cell with a positive quantity -- the unit
    _resolve_column and the base-card lookup operate on."""

    row_number: int
    raw_set: str | None
    raw_number: str | None
    column: str
    quantity: int


def _read_header(sheet) -> dict[str, int]:
    """Maps column name -> 0-based cell index from the sheet's first row.
    Blank header cells are skipped (openpyxl pads short rows with None)."""
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return {
        str(name).strip(): idx
        for idx, name in enumerate(header_row)
        if name is not None and str(name).strip()
    }


def _cell_quantity(value: object) -> int | None:
    """§2.3/§10 (owner-locked): positive-only semantics -- blank/None
    asserts nothing (never treated as zero/removal). A literal 0, should
    one ever appear (none do in the owner's real export), is likewise
    ignored as no-claim -- returns None here exactly like a blank cell, so
    the melt step never emits a candidate for it. A non-numeric or
    negative cell is malformed and also ignored (no candidate) rather than
    raising -- this format has no per-cell malformed-row report slot the
    way canonical CSV/JSON quantity does, since a "row" here is the whole
    spreadsheet line, not one cell."""
    if value is None:
        return None
    try:
        quantity = int(value)
    except (TypeError, ValueError):
        return None
    return quantity if quantity > 0 else None


def parse_swunlimiteddb_xlsx(raw_bytes: bytes, db: Session) -> ParseResult:
    """§2/§5/§10. Reads the "Data" sheet from raw XLSX bytes (never
    text-decoded -- the router's format-detection seam routes here BEFORE
    any utf-8 decode attempt, since this is a binary format). Refuses
    outright (never best-effort parsed) when the file isn't a readable
    workbook, the "Data" sheet is missing, or the header lacks Set/Base
    card id -- same posture as parse_swudb_csv. Row-level resolution
    failures (unmapped set, unknown base card, unmapped/ambiguous column)
    are NOT refused -- they surface as unresolved/ambiguous report rows
    exactly like any other import problem row."""
    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(raw_bytes), read_only=True, data_only=True
        )
    except Exception as exc:  # noqa: BLE001 -- openpyxl's own exception
        # surface (zipfile.BadZipFile, KeyError, etc.) varies by corruption
        # mode; this module's documented parse-failure surface is
        # UnparseableFileError only, matching inventory_io._decode's own
        # "never leak the underlying library's exception type" contract.
        raise UnparseableFileError(f"not a readable XLSX workbook: {exc}") from exc

    if DATA_SHEET_NAME not in workbook.sheetnames:
        raise UnparseableFileError(
            f"no {DATA_SHEET_NAME!r} sheet in workbook "
            f"(found: {', '.join(workbook.sheetnames)})"
        )
    sheet = workbook[DATA_SHEET_NAME]

    header = _read_header(sheet)
    if not REQUIRED_COLUMNS.issubset(header):
        raise UnparseableFileError(
            "not a recognizable sw-unlimited-db export (need Set, Base card id)"
        )

    unrecognized = [name for name in header if name not in KNOWN_COLUMNS]
    quantity_columns_present = [c for c in QUANTITY_COLUMNS if c in header]

    set_idx = header["Set"]
    number_idx = header["Base card id"]

    candidates: list[_Candidate] = []
    row_number = 0
    for raw_row in sheet.iter_rows(min_row=2, values_only=True):
        row_number += 1
        raw_set = raw_row[set_idx] if set_idx < len(raw_row) else None
        raw_number = raw_row[number_idx] if number_idx < len(raw_row) else None
        raw_set = str(raw_set).strip() if raw_set is not None else None
        raw_number = str(raw_number).strip() if raw_number is not None else None
        if not raw_set or not raw_number:
            # §10: an unusable row identity (blank Set/Base card id) has no
            # positive-only claim to make regardless of quantity cells --
            # skip melting it. This format has no malformed_row concept
            # (see _cell_quantity's docstring); a garbage row simply melts
            # to zero candidates.
            continue
        for column in quantity_columns_present:
            idx = header[column]
            value = raw_row[idx] if idx < len(raw_row) else None
            quantity = _cell_quantity(value)
            if quantity is None:
                continue
            candidates.append(
                _Candidate(
                    row_number=row_number,
                    raw_set=raw_set,
                    raw_number=raw_number,
                    column=column,
                    quantity=quantity,
                )
            )

    workbook.close()

    # §5 steps 1-2 / §10: map every distinct (raw Set, raw Base card id) up
    # front, so the base-card-first bulk lookup runs ONE query regardless
    # of file size -- mirrors swudb_import.parse_swudb_csv's own bulk-map-
    # then-lookup shape. BL-199: pairs are partitioned by tokenness (the
    # T- prefix _normalize_number strips is the row's only token signal --
    # see _is_token_id) and fetched in two token-split calls, so a token
    # row (T1) and a real-card row (1) sharing the same bare number never
    # see each other's base cards.
    pairs_by_tokenness: dict[bool, set[tuple[str, str]]] = {False: set(), True: set()}
    mapped_by_candidate: list[tuple[tuple[str, str], bool] | None] = []
    for cand in candidates:
        set_code = SET_CODE_MAP.get(cand.raw_set.upper())
        if set_code is None:
            mapped_by_candidate.append(None)
            continue
        number = _normalize_number(cand.raw_number)
        is_token = _is_token_id(cand.raw_number)
        pair = (set_code, number)
        mapped_by_candidate.append((pair, is_token))
        pairs_by_tokenness[is_token].add(pair)

    base_card_variants_by_tokenness = {
        tokens: inventory_import_repo.get_base_card_variants_by_set_and_number(
            db, list(token_pairs), tokens=tokens
        )
        for tokens, token_pairs in pairs_by_tokenness.items()
    }

    parsed_rows: list[ParsedRow] = []
    # BL-186: every melted candidate becomes its own report row -- unlike
    # canonical/SWUDB parsers, row_number here is a synthetic per-candidate
    # sequence (a single spreadsheet row can melt into several candidates,
    # §5.1), not the spreadsheet's own row index, so every candidate gets a
    # distinct, stable ordinal.
    for ordinal, (cand, mapped) in enumerate(
        zip(candidates, mapped_by_candidate, strict=True), start=1
    ):
        if cand.column in UNMAPPED_COLUMNS:
            # §6/§10: unmapped-column check runs before set/base-card
            # resolution -- a column with no confident family mapping is a
            # problem row regardless of whether its set/base-card would
            # otherwise resolve cleanly. variant_type is deliberately left
            # None (never the column name) -- same reason as swudb_import's
            # own preresolved rows (module docstring there): keeps
            # _triple_of(row) always None for a preresolved row, so (a)
            # compute_import's own bulk triple lookup never wastes a query
            # slot on a synthetic (set, number, column) triple that could
            # never match a real variant_type, and (b) inventory_io's
            # _merge_duplicates never silently sums two DISTINCT problem
            # rows into one just because they share a raw identity -- every
            # unresolved/ambiguous row stays individually visible in the
            # report, matching SWUDB's established transparency posture.
            parsed_rows.append(
                ParsedRow(
                    row_number=ordinal,
                    swuapi_uuid=None,
                    set_code=cand.raw_set,
                    card_number=cand.raw_number,
                    variant_type=None,
                    quantity=cand.quantity,
                    malformed_quantity=False,
                    preresolved_reason="unmapped_column",
                )
            )
            continue

        if mapped is None:
            # §3/§10 (owner-locked): unmapped set (incl. HMW and anything
            # else outside the 10 base sets) -- never guessed, same
            # generic reason SWUDB's own unmapped-set case uses.
            parsed_rows.append(
                ParsedRow(
                    row_number=ordinal,
                    swuapi_uuid=None,
                    set_code=cand.raw_set,
                    card_number=cand.raw_number,
                    variant_type=None,
                    quantity=cand.quantity,
                    malformed_quantity=False,
                    preresolved_reason="unmapped_set",
                )
            )
            continue

        pair, is_token = mapped
        mapped_set, number = pair
        variants = base_card_variants_by_tokenness[is_token].get(pair, [])
        outcome = _resolve_column(variants, cand.column, mapped_set)
        parsed_rows.append(
            ParsedRow(
                row_number=ordinal,
                swuapi_uuid=outcome.swuapi_uuid,
                set_code=None if outcome.swuapi_uuid else mapped_set,
                card_number=None if outcome.swuapi_uuid else number,
                variant_type=None,
                quantity=cand.quantity,
                malformed_quantity=False,
                preresolved_reason=outcome.reason,
                preresolved_candidates=outcome.candidates or None,
            )
        )

    notes = ParseNotes(unrecognized_columns=unrecognized)
    # BL-186: reuses inventory_io's own duplicate-identity fold, same as
    # swudb_import -- a resolved candidate already carries a swuapi_uuid by
    # this point, so two melted candidates resolving to the same variant
    # (e.g. the same base card entered twice, or two columns resolving to
    # the same variant -- not expected in practice but not precluded)
    # get summed exactly like two canonical rows sharing a raw uuid would.
    rows = _merge_duplicates(parsed_rows, notes)
    return ParseResult(rows=rows, notes=notes)
